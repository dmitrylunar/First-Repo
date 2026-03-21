from database import get_connection
from models import Client
from utils import consolidate
from validators import valid_name, valid_number, valid_email, valid_id
from validators import is_valid_name, is_valid_number, is_valid_email
import datetime
import smtplib
import openpyxl
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
load_dotenv()

HEADER_MAP = {
    'name': 'name', 'имя': 'name',
    'email': 'email', 'почта': 'email', 'e-mail': 'email',
    'number': 'number', 'номер': 'number', 'телефон': 'number'
}

# ── Клиенты ────────────────────────────────────────────────────────────────

def add_client():
    print("Дата записи: ", datetime.date.today())
    print("Введите данные о клиенте...")
    name = valid_name(input("Имя клиента: "))
    number = valid_number(input("Номер клиента: "))
    email = valid_email(input("Емейл клиента: "))
    return Client(name, number, email)

def save_client(client):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute('INSERT INTO clients (name, number, email) VALUES (%s, %s, %s)',
                (client.name, client.number, client.email))
    conn.commit()
    cur.close()
    conn.close()
    print("Клиент сохранен!")

def get_all_clients():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM clients')
    clients = cur.fetchall()
    cur.close()
    conn.close()
    return clients

def client_list():
    for person in get_all_clients():
        print(f"ID: {person[0]} \nИмя: {person[1]}\nНомер: {person[2]}\nПочта: {person[3]}\n")

def find_client(search_info):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""SELECT * FROM clients
                   WHERE CAST(id AS text) LIKE %s
                      OR name LIKE %s
                      OR number LIKE %s
                      OR email LIKE %s
                """, (f"%{search_info}%",) * 4)
    found = cur.fetchall()
    cur.close()
    conn.close()
    return found

@consolidate
def delete_client(choose):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM clients WHERE id = %s', (choose,))
    conn.commit()
    cur.close()
    conn.close()
    print("Клиент удален...")

def delete_finding_client():
    client_list()
    print('Выберете ID клиента которого хотите удалить..')
    choose = valid_id(input("Ввод: "))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM clients WHERE id = %s', (choose,))
    client_to_delete = cur.fetchone()
    cur.close()
    conn.close()
    if client_to_delete is None:
        print("Клиент не обнаружен")
    else:
        print(f"ID: {client_to_delete[0]} \nИмя: {client_to_delete[1]}\nНомер: {client_to_delete[2]}\nПочта: {client_to_delete[3]}\n")
        delete_client(choose)

# ── Базы данных ────────────────────────────────────────────────────────────

def get_databases_list():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT db_id, db_name, COUNT(*) as total
        FROM clients
        WHERE db_id IS NOT NULL
        GROUP BY db_id, db_name
        ORDER BY db_id
    """)
    databases = cur.fetchall()
    cur.close()
    conn.close()
    return databases

def load_database_from_excel(file_path, db_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    raw_headers = [str(cell.value).lower().strip() for cell in ws[1]]
    headers = [HEADER_MAP.get(h, h) for h in raw_headers]

    conn = get_connection()
    cur = conn.cursor()
    cur.execute('SELECT COALESCE(MAX(db_id), 0) + 1 FROM clients')
    db_id = cur.fetchone()[0]

    added = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        row_dict = dict(zip(headers, row))
        name = row_dict.get('name')
        number = row_dict.get('number')
        email = row_dict.get('email')

        if not (is_valid_name(name) and is_valid_number(number) and is_valid_email(email)):
            skipped += 1
            continue

        number_str = str(int(float(str(number))))

        cur.execute(
            'INSERT INTO clients (name, number, email, db_id, db_name) VALUES (%s, %s, %s, %s, %s)',
            (str(name), number_str, str(email), db_id, db_name)
        )
        added += 1

    conn.commit()
    cur.close()
    conn.close()
    return added, skipped

def delete_database(db_name):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM clients WHERE db_name = %s', (db_name,))
    count = cur.fetchone()[0]
    if count == 0:
        cur.close()
        conn.close()
        return 0
    cur.execute('DELETE FROM clients WHERE db_name = %s', (db_name,))
    conn.commit()
    cur.close()
    conn.close()
    return count

# ── Email ──────────────────────────────────────────────────────────────────

def build_email(from_email, to_email, to_name, subject, message):
    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    body = f"Привет, {to_name}!\n\n{message}"
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    return msg

def send_email(to_email, to_name, subject, message):
    from_email = os.getenv('GMAIL_USER')
    password = os.getenv('GMAIL_PASSWORD')
    msg = build_email(from_email, to_email, to_name, subject, message)
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(from_email, password)
            server.send_message(msg)
        print(f"Отправлено: {to_name} <{to_email}>")
        return True
    except Exception as e:
        print(f"Ошибка для {to_email}: {e}")
        return False

def send_bulk_email(subject, message, db_id=None):
    conn = get_connection()
    cur = conn.cursor()
    if db_id is not None:
        cur.execute('SELECT id, name, email FROM clients WHERE db_id = %s', (db_id,))
    else:
        cur.execute('SELECT id, name, email FROM clients')
    clients = cur.fetchall()
    cur.close()
    conn.close()

    if not clients:
        print("База клиентов пуста")
        return 0, 0

    print(f"\nНачинаем рассылку для {len(clients)} клиентов...\n")

    success = 0
    failed = 0

    for client_id, name, email in clients:
        if email:
            result = send_email(email, name, subject, message)
            if result:
                success += 1
            else:
                failed += 1
        else:
            print(f"Нет email у клиента {name} (id={client_id})")

    print(f"\nГотово: {success} отправлено, {failed} ошибок")
    return success, failed